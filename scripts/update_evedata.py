#!/usr/bin/env python3
"""
Download evedata.xlsx from ceve-market.org and convert to JSON lookup files.
Output:
  - data/eve_names.json          : {id: "中文名", ...} for all entity types
  - data/eve_station_systems.json: {stationId: systemId, ...} for station→system mapping
  - data/evedata_meta.json       : update metadata
"""
import json
import os
import sys
import urllib.request
from datetime import datetime

import openpyxl

XLSX_URL = 'https://www.ceve-market.org/dumps/evedata.xlsx'
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, 'data')
XLSX_PATH = os.path.join(DATA_DIR, 'evedata.xlsx')

def main():
    os.makedirs(DATA_DIR, exist_ok=True)

    # 1. Download xlsx
    print(f"Downloading {XLSX_URL} ...")
    urllib.request.urlretrieve(XLSX_URL, XLSX_PATH)
    file_size = os.path.getsize(XLSX_PATH)
    print(f"  Downloaded: {file_size:,} bytes")

    # 2. Parse xlsx
    print("Parsing xlsx ...")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    names = {}           # unified id -> name
    station_systems = {} # stationId -> systemId
    counts = {}

    # Sheet 1: 物品列表 (typeID, 物品名称, ...)
    ws = wb['物品列表']
    c = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1]:
            names[int(row[0])] = str(row[1])
            c += 1
    counts['types'] = c

    # Sheet 2: 星域列表 (星域ID, 星域名字)
    ws = wb['星域列表']
    c = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1]:
            names[int(row[0])] = str(row[1])
            c += 1
    counts['regions'] = c

    # Sheet 3: 星座列表 (星座ID, 星座名字, 星域ID, 星域名字)
    ws = wb['星座列表']
    c = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1]:
            names[int(row[0])] = str(row[1])
            c += 1
    counts['constellations'] = c

    # Sheet 4: 星系列表 (星系ID, 星系名字, ...)
    ws = wb['星系列表']
    c = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1]:
            names[int(row[0])] = str(row[1])
            c += 1
    counts['systems'] = c

    # Sheet 5: NPC空间站 (空间站ID, 空间站名称, 星系ID, ...)
    ws = wb['NPC空间站']
    c = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1]:
            sid = int(row[0])
            names[sid] = str(row[1])
            if row[2] is not None:
                station_systems[sid] = int(row[2])
            c += 1
    counts['npc_stations'] = c

    # Sheet 6: 玩家公开建筑 (建筑物ID, 建筑物名称, 建筑物类型, 星系ID, ...)
    # Sheet name contains date, find it dynamically
    structure_sheet = [s for s in wb.sheetnames if '玩家' in s or '建筑' in s]
    if structure_sheet:
        ws = wb[structure_sheet[0]]
        c = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and row[1]:
                sid = int(row[0])
                names[sid] = str(row[1])
                if row[3] is not None:
                    station_systems[sid] = int(row[3])
                c += 1
        counts['player_structures'] = c

    wb.close()

    # 3. Write JSON files
    # eve_names.json: keys must be strings in JSON
    names_out = {str(k): v for k, v in names.items()}
    names_path = os.path.join(DATA_DIR, 'eve_names.json')
    with open(names_path, 'w', encoding='utf-8') as f:
        json.dump(names_out, f, ensure_ascii=False, separators=(',', ':'))
    print(f"  eve_names.json: {len(names):,} entries, {os.path.getsize(names_path):,} bytes")

    # eve_station_systems.json
    ss_out = {str(k): v for k, v in station_systems.items()}
    ss_path = os.path.join(DATA_DIR, 'eve_station_systems.json')
    with open(ss_path, 'w', encoding='utf-8') as f:
        json.dump(ss_out, f, separators=(',', ':'))
    print(f"  eve_station_systems.json: {len(station_systems):,} entries, {os.path.getsize(ss_path):,} bytes")

    # evedata_meta.json
    meta = {
        'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'source': XLSX_URL,
        'counts': counts,
        'total_names': len(names),
    }
    meta_path = os.path.join(DATA_DIR, 'evedata_meta.json')
    with open(meta_path, 'w', encoding='utf-8') as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
    print(f"  evedata_meta.json written")

    print(f"\nDone! Total entries: {len(names):,}")
    for k, v in counts.items():
        print(f"  {k}: {v:,}")

if __name__ == '__main__':
    main()
